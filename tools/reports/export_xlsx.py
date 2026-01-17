"""
XLSX Reports Exporter (read-only)

MontrixBot v2.3.5
- Dependency-hardened (openpyxl)
- UX-friendly CLI
- Canon-safe: NO runtime writes, read-only sources only
"""

import sys
import json
import csv
import argparse
from pathlib import Path
from datetime import datetime


# -------- Dependency hardening --------
try:
    # openpyxl is optional; if present, use it for XLSX generation
    from openpyxl import Workbook  # type: ignore
    _OPENPYXL_AVAILABLE = True
except ImportError:
    # do not abort if openpyxl is missing; fall back to CSV exports
    _OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore


RUNTIME_DIR = Path("runtime")
EXPORTS_DIR = Path("exports")


def _load_json(path: Path) -> dict:
    try:
        if path.exists():
            with path.open("r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _load_jsonl(path: Path) -> list[dict]:
    rows = []
    try:
        if path.exists():
            with path.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rows.append(json.loads(line))
                    except Exception:
                        continue
    except Exception:
        pass
    return rows


def _load_equity_history(path: Path) -> list[tuple]:
    rows = []
    try:
        if path.exists():
            with path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    try:
                        ts = float(r.get("ts"))
                        eq = float(r.get("equity"))
                        rows.append((ts, eq))
                    except Exception:
                        continue
    except Exception:
        pass
    return rows


def export_xlsx(out_path: Path) -> Path:
    """
    Export portfolio data to XLSX if openpyxl is installed;
    otherwise fall back to CSV files. Returns the path to the
    generated file or directory.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Load runtime data
    status = _load_json(RUNTIME_DIR / "status.json")
    sim_state = _load_json(RUNTIME_DIR / "sim_state.json")
    trades = _load_jsonl(RUNTIME_DIR / "trades.jsonl")
    equity = _load_equity_history(RUNTIME_DIR / "equity_history.csv")

    # Prepare summary rows
    summary_rows = [
        ("equity_current", status.get("equity") or sim_state.get("equity")),
        ("pnl_day_pct", status.get("pnl_day_pct")),
        ("pnl_total_pct", status.get("pnl_total_pct")),
        ("open_positions", status.get("open_positions_count")),
        ("mode", status.get("mode")),
        ("snapshot_ts", status.get("snapshot_ts") or status.get("ts")),
    ]

    if _OPENPYXL_AVAILABLE and Workbook is not None:
        # Generate an XLSX workbook using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(("field", "value"))
        for k, v in summary_rows:
            ws.append((k, v))
        ws_eq = wb.create_sheet("Equity")
        ws_eq.append(("ts", "equity"))
        for ts, eq in equity:
            ws_eq.append((ts, eq))
        ws_tr = wb.create_sheet("Trades")
        ws_tr.append(("ts", "symbol", "side", "qty", "price", "status", "source"))
        for t in trades:
            ws_tr.append((
                t.get("ts"),
                t.get("symbol"),
                t.get("side"),
                t.get("qty"),
                t.get("price"),
                t.get("status"),
                t.get("source"),
            ))
        wb.save(out_path)
        return out_path
    else:
        # Fall back to CSV exports: write summary, equity and trades as separate CSV files
        base_stem = out_path.stem
        summary_path = EXPORTS_DIR / f"{base_stem}_summary.csv"
        equity_path = EXPORTS_DIR / f"{base_stem}_equity.csv"
        trades_path = EXPORTS_DIR / f"{base_stem}_trades.csv"
        try:
            with summary_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["field", "value"])
                for row in summary_rows:
                    writer.writerow(row)
        except Exception:
            pass
        try:
            with equity_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["ts", "equity"])
                for ts, eq in equity:
                    writer.writerow([ts, eq])
        except Exception:
            pass
        try:
            with trades_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["ts", "symbol", "side", "qty", "price", "status", "source"])
                for t in trades:
                    writer.writerow([
                        t.get("ts"),
                        t.get("symbol"),
                        t.get("side"),
                        t.get("qty"),
                        t.get("price"),
                        t.get("status"),
                        t.get("source"),
                    ])
        except Exception:
            pass
        return EXPORTS_DIR


def main():
    parser = argparse.ArgumentParser(description="Export XLSX reports (read-only)")
    parser.add_argument(
        "--out",
        type=str,
        help="Output .xlsx file or directory (default: exports/portfolio_report_<date>.xlsx)",
    )
    args = parser.parse_args()

    if args.out:
        out = Path(args.out)
        if out.suffix.lower() != ".xlsx":
            out = out / f"portfolio_report_{datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    else:
        out = EXPORTS_DIR / f"portfolio_report_{datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)

    result = export_xlsx(out)
    print(f"XLSX exported to: {result}")


if __name__ == "__main__":
    main()
